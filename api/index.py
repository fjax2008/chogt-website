"""
桶号管理 API — Vercel Serverless Function
直接连接 Turso 数据库，无 Render 依赖
"""
import json
import asyncio
import os
import io
from datetime import datetime, timezone, timedelta
from urllib.parse import urlparse, parse_qs, unquote

# ─── Turso 云数据库 ───────────────────────────────────────
TURSO_URL = os.environ.get("TURSO_URL", "https://barrel-db-fjax2008.aws-ap-south-1.turso.io")
TURSO_TOKEN = os.environ.get("TURSO_TOKEN", "eyJhbGciOiJFZERTQSIsInR5cCI6IkpXVCJ9.eyJpYXQiOjE3ODEyNDk3NDQsImlkIjoiMDE5ZWJhYzAtODkzYy03ZjYyLWIyMjUtMDU3ZTdiZGYyYzk0In0.D0MzszPVPj49eKhbGd3xLyYKgLx5M3VRsBpZFqNNffxrYQp7LJCx5JfjSQFgBMP3IOLQ3VoTMRoBikmMCZNcCA")
TZ = timezone(timedelta(hours=7))

def now_str(fmt="%Y-%m-%d %H:%M:%S"):
    return datetime.now(TZ).strftime(fmt)


# ─── 数据库操作（兼容 Vercel Functions 事件循环）──────────
def _run_async(coro):
    """在 Vercel Functions 中安全执行异步操作"""
    try:
        return asyncio.run(coro)
    except RuntimeError:
        try:
            loop = asyncio.get_event_loop()
            if loop.is_running():
                import nest_asyncio
                nest_asyncio.apply()
                return asyncio.run(coro)
            return loop.run_until_complete(coro)
        except:
            import nest_asyncio
            nest_asyncio.apply()
            return asyncio.run(coro)


def db(sql, params=None):
    from libsql_client import create_client
    async def _exec():
        client = create_client(url=TURSO_URL, auth_token=TURSO_TOKEN)
        try:
            rs = await client.execute(sql, params or [])
            return [dict(zip(rs.columns, row)) for row in rs.rows]
        finally:
            await client.close()
    return _run_async(_exec())


def db_batch(stmts):
    from libsql_client import create_client
    async def _exec():
        client = create_client(url=TURSO_URL, auth_token=TURSO_TOKEN)
        try:
            return await client.batch(stmts)
        finally:
            await client.close()
    return _run_async(_exec())


# ─── 响应工具 ─────────────────────────────────────────────
def json_response(data, status=200):
    return {
        "statusCode": status,
        "headers": {
            "Content-Type": "application/json; charset=utf-8",
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, POST, DELETE, OPTIONS",
            "Access-Control-Allow-Headers": "Content-Type",
        },
        "body": json.dumps(data, ensure_ascii=False)
    }


def excel_response(fileobj, filename):
    return {
        "statusCode": 200,
        "headers": {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Allow-Origin": "*",
        },
        "body": fileobj.getvalue().hex(),
        "isBase64Encoded": True,
    }


# ─── 路由 ─────────────────────────────────────────────────
def handler(event, context):
    # Vercel Python runtime: event is the request
    method = (event.get("httpMethod") or event.get("method") or "GET").upper()
    path = event.get("rawPath") or event.get("path") or "/api"
    body = event.get("body") or "{}"
    query = event.get("queryStringParameters") or {}

    # CORS preflight
    if method == "OPTIONS":
        return {"statusCode": 200, "headers": {"Access-Control-Allow-Origin": "*", "Access-Control-Allow-Methods": "GET, POST, DELETE, OPTIONS", "Access-Control-Allow-Headers": "Content-Type"}, "body": ""}

    # Parse body if it's base64
    if event.get("isBase64Encoded") and body:
        import base64
        body = base64.b64decode(body).decode("utf-8")

    try:
        req_data = json.loads(body) if body else {}
    except:
        req_data = {}

    # ── GET /api/inventory ──
    if method == "GET" and path in ("/api/inventory", "/api/inventory/"):
        rows = db("SELECT * FROM barrel_inventory WHERE status='in' ORDER BY location")
        return json_response(rows)

    # ── GET /api/inventory/{keyword} ──
    if method == "GET" and path.startswith("/api/inventory/") and not path.startswith("/api/inventory/location/"):
        keyword = unquote(path.split("/api/inventory/")[1].rstrip("/"))
        rows = db("SELECT * FROM barrel_inventory WHERE barrel_no LIKE ? ORDER BY status='in' DESC, update_time DESC",
                  [f"%{keyword}%"])
        if not rows:
            rows = db("SELECT * FROM barrel_inventory WHERE location LIKE ? AND status='in' ORDER BY location",
                      [f"%{keyword}%"])
        return json_response(rows)

    # ── GET /api/inventory/location/{location} ──
    if method == "GET" and path.startswith("/api/inventory/location/"):
        location = unquote(path.split("/api/inventory/location/")[1].rstrip("/"))
        rows = db("SELECT * FROM barrel_inventory WHERE location=? AND status='in'", [location])
        return json_response(rows)

    # ── POST /api/inbound/batch ──
    if method == "POST" and path in ("/api/inbound/batch", "/api/inbound/batch/"):
        barrel_nos = req_data.get("barrel_nos", [])
        location = req_data.get("location", "")
        if not barrel_nos or not location:
            return json_response({"success": False, "msg": "缺少参数"}, 400)
        now = now_str()
        # 批量查询已有桶号（一趟往返）
        clean = [bn.strip() for bn in barrel_nos if bn.strip()]
        if not clean:
            return json_response({"success": True, "location": location, "success_count": 0, "success_list": []})
        placeholders = ",".join(["?"] * len(clean))
        existing = db(f"SELECT barrel_no FROM barrel_inventory WHERE barrel_no IN ({placeholders})", clean)
        exists_set = {r["barrel_no"] for r in existing}
        # 批量写入（一趟往返）
        stmts = []
        for bn in clean:
            action = "移动" if bn in exists_set else "入库"
            stmts.append(("INSERT OR REPLACE INTO barrel_inventory (barrel_no, location, status, update_time) VALUES (?,?,'in',?)",
                          [bn, location, now]))
            stmts.append(("INSERT INTO barrel_log (barrel_no, action, location, created_at) VALUES (?,?,?,?)",
                          [bn, action, location, now]))
        db_batch(stmts)
        return json_response({"success": True, "location": location, "success_count": len(clean), "success_list": clean})

    # ── DELETE /api/barrel/{barrel_no} ──
    if method == "DELETE" and path.startswith("/api/barrel/"):
        barrel_no = unquote(path.split("/api/barrel/")[1].rstrip("/"))
        if not barrel_no:
            return json_response({"success": False, "msg": "缺少桶号"}, 400)
        rows = db("SELECT * FROM barrel_inventory WHERE barrel_no=?", [barrel_no])
        if not rows or rows[0]["status"] != "in":
            return json_response({"success": False, "msg": f"桶号 {barrel_no} 不存在"}, 404)
        now = now_str()
        loc = rows[0]["location"]
        db_batch([
            ("DELETE FROM barrel_inventory WHERE barrel_no=?", [barrel_no]),
            ("INSERT INTO barrel_log (barrel_no, action, location, created_at) VALUES (?,?,?,?)",
             [barrel_no, "删除", loc, now]),
        ])
        return json_response({"success": True, "msg": f"桶号 {barrel_no}（in / {loc}）已删除"})

    # ── DELETE /api/barrel/delete/batch ──

    if method == "DELETE" and path in ("/api/barrel/delete/batch", "/api/barrel/delete/batch/"):
        barrel_nos = req_data.get("barrel_nos", [])
        if not barrel_nos:
            return json_response({"success": False, "msg": "缺少桶号"}, 400)
        now = now_str()
        deleted_list = []
        for bn in barrel_nos:
            bn = bn.strip()
            if not bn:
                continue
            rows = db("SELECT * FROM barrel_inventory WHERE barrel_no=?", [bn])
            if rows and rows[0]["status"] == "in":
                loc = rows[0]["location"]
                db_batch([
                    ("DELETE FROM barrel_inventory WHERE barrel_no=?", [bn]),
                    ("INSERT INTO barrel_log (barrel_no, action, location, created_at) VALUES (?,?,?,?)",
                     [bn, "删除", loc, now]),
                ])
                deleted_list.append(bn)
        return json_response({"success": True, "deleted_count": len(deleted_list), "deleted_list": deleted_list})

    # ── GET /api/logs ──
    if method == "GET" and path in ("/api/logs", "/api/logs/"):
        limit = int(query.get("limit", 100))
        rows = db("SELECT * FROM barrel_log ORDER BY id DESC LIMIT ?", [limit])
        return json_response(rows)

    # ── GET /api/export ──
    if method == "GET" and path in ("/api/export", "/api/export/"):
        rows = db("SELECT * FROM barrel_inventory WHERE status='in' ORDER BY location")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = "桶号库存"
            header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
            headers = ["桶号", "储位", "状态", "更新时间"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            data_font = Font(name="微软雅黑", size=10)
            data_align = Alignment(horizontal="center", vertical="center")
            for i, row in enumerate(rows, 2):
                for j, key in enumerate(["barrel_no", "location", "status", "update_time"], 1):
                    cell = ws.cell(row=i, column=j, value=row.get(key, ""))
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = thin_border
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 22
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            import base64
            return {
                "statusCode": 200,
                "headers": {
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "Content-Disposition": "attachment; filename=barrel_inventory.xlsx",
                    "Access-Control-Allow-Origin": "*",
                },
                "body": base64.b64encode(buf.getvalue()).decode(),
                "isBase64Encoded": True,
            }
        except Exception as e:
            return json_response({"error": str(e)}, 500)

    return json_response({"error": "Not Found"}, 404)
